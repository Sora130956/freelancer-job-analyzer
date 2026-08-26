"""Tests for the Excel generator."""
from io import BytesIO

import openpyxl

from backend.services.data_processor import BUDGET_BINS, enrich_projects
from backend.services.excel_generator import (
    BUDGET_SHEET,
    PROJECT_COLUMNS,
    PROJECTS_SHEET,
    SKILLS_SHEET,
    generate_excel,
)

RATES = {"USD": 1.0, "EUR": 1.1}


def _project(
    project_id=1,
    minimum=50.0,
    maximum=150.0,
    code="USD",
    jobs=None,
    bid_avg=100.0,
    bid_count=12,
    seo_url=None,
):
    return {
        "id": project_id,
        "title": f"Project {project_id}",
        "seo_url": seo_url or f"python/project-{project_id}",
        "type": "fixed",
        "time_submitted": 1700000000,
        "budget": {"minimum": minimum, "maximum": maximum},
        "currency": {"code": code, "exchange_rate": RATES[code]},
        "jobs": jobs if jobs is not None else [{"id": 3, "name": "Python"}],
        "bid_stats": {"bid_count": bid_count, "bid_avg": bid_avg},
    }


def _workbook(projects):
    """Run the generator and load the produced bytes back with openpyxl."""
    return openpyxl.load_workbook(BytesIO(generate_excel(projects)))


def _rows(sheet):
    return list(sheet.iter_rows(values_only=True))


def test_workbook_has_three_sheets_in_order():
    """AC-005 requires exactly these three sheets in this order."""
    wb = _workbook(enrich_projects([_project()], RATES))

    assert wb.sheetnames == [PROJECTS_SHEET, SKILLS_SHEET, BUDGET_SHEET]


def test_projects_sheet_header_matches_acceptance_columns():
    """Column names are part of the deliverable contract."""
    wb = _workbook(enrich_projects([_project()], RATES))

    assert _rows(wb[PROJECTS_SHEET])[0] == tuple(PROJECT_COLUMNS)
    assert PROJECT_COLUMNS == [
        "Title",
        "Skills",
        "Budget (USD)",
        "Avg Bid",
        "Bid Count",
        "Type",
        "Posted",
        "URL",
    ]


def test_projects_sheet_has_one_row_per_project():
    """Row count must be project count + 1 header row."""
    projects = enrich_projects([_project(project_id=i) for i in range(1, 6)], RATES)

    wb = _workbook(projects)

    assert len(_rows(wb[PROJECTS_SHEET])) == 6


def test_projects_row_carries_converted_amounts_and_joined_skills():
    """Amounts come from the USD fields; skills collapse into one cell."""
    project = _project(
        code="EUR",
        maximum=150.0,
        bid_avg=100.0,
        jobs=[{"id": 3, "name": "Python"}, {"id": 7, "name": "Scrapy"}],
    )

    row = _rows(_workbook(enrich_projects([project], RATES))[PROJECTS_SHEET])[1]

    assert row[0] == "Project 1"
    assert row[1] == "Python, Scrapy"
    assert row[2] == 165.0
    assert row[3] == 110.0
    assert row[4] == 12
    assert row[5] == "fixed"


def test_projects_row_builds_absolute_url_from_seo_url():
    """The URL column must be clickable, not a bare slug."""
    project = _project(seo_url="python/build-a-scraper")

    row = _rows(_workbook(enrich_projects([project], RATES))[PROJECTS_SHEET])[1]

    assert row[7] == "https://www.freelancer.com/projects/python/build-a-scraper"


def test_projects_row_formats_posted_timestamp_as_utc_text():
    """The epoch seconds from the API must become a readable UTC stamp."""
    row = _rows(_workbook(enrich_projects([_project()], RATES))[PROJECTS_SHEET])[1]

    assert row[6] == "2023-11-14 22:13"


def test_projects_row_leaves_missing_amounts_empty():
    """Missing budget/bid data must stay blank instead of becoming 0."""
    projects = enrich_projects([_project()], RATES)
    projects[0]["budget_max_usd"] = None
    projects[0]["bid_avg_usd"] = None

    row = _rows(_workbook(projects)[PROJECTS_SHEET])[1]

    assert row[2] is None
    assert row[3] is None


def test_skills_sheet_lists_every_skill_descending():
    """AC-005: the skills sheet is the full frequency table, not just top 10."""
    projects = enrich_projects(
        [
            _project(project_id=i, jobs=[{"id": i, "name": f"Skill{i}"}] * (14 - i))
            for i in range(1, 14)
        ],
        RATES,
    )

    rows = _rows(_workbook(projects)[SKILLS_SHEET])

    assert rows[0] == ("Skill name", "Count")
    assert len(rows) == 14
    counts = [row[1] for row in rows[1:]]
    assert counts == sorted(counts, reverse=True)


def test_budget_sheet_lists_all_bins_in_fixed_order():
    """Every bin appears, even at zero, so the report layout is stable."""
    projects = enrich_projects([_project(maximum=300.0)], RATES)

    rows = _rows(_workbook(projects)[BUDGET_SHEET])

    assert rows[0] == ("Bin", "Count")
    assert [row[0] for row in rows[1:]] == [label for label, _, _ in BUDGET_BINS]
    assert dict(rows[1:])["$150-$500"] == 1


def test_empty_result_still_produces_three_sheets_with_headers():
    """AC-005 edge case: an empty search must produce a valid workbook."""
    wb = _workbook([])

    assert wb.sheetnames == [PROJECTS_SHEET, SKILLS_SHEET, BUDGET_SHEET]
    assert _rows(wb[PROJECTS_SHEET]) == [tuple(PROJECT_COLUMNS)]
    assert _rows(wb[SKILLS_SHEET]) == [("Skill name", "Count")]
    assert len(_rows(wb[BUDGET_SHEET])) == len(BUDGET_BINS) + 1


def test_generate_excel_returns_xlsx_bytes():
    """The route layer streams bytes, so the entry point must return bytes."""
    payload = generate_excel(enrich_projects([_project()], RATES))

    assert isinstance(payload, bytes)
    assert payload[:2] == b"PK"
